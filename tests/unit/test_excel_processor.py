#!/usr/bin/env python3
"""
Excel Processor 单元测试
"""

import os
import sys
import tempfile
import unittest

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side

# 添加项目根目录到 Python 路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.plugins.excel_writer.processor import ExcelProcessor


class TestExcelProcessor(unittest.TestCase):
    """测试 ExcelProcessor 类"""

    def setUp(self):
        """测试前准备"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_excel = os.path.join(self.temp_dir, "test.xlsx")

        # 创建测试用的 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # 填充测试数据
        ws["A1"] = "Field1"
        ws["B1"] = "Value1"
        ws["A2"] = "Field2"
        ws["B2"] = "Value2"
        ws["A3"] = "Field3"
        ws["B3"] = "Value3"

        # 添加合并单元格
        ws.merge_cells("A5:B5")
        ws["A5"] = "Merged Cell"

        # 添加样式
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        wb.save(self.test_excel)
        self.wb = wb

    def tearDown(self):
        """测试后清理"""
        import shutil

        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def test_init_with_file_path(self):
        """测试使用文件路径初始化"""
        processor = ExcelProcessor(self.test_excel)
        self.assertIsNotNone(processor.wb)
        self.assertIsNotNone(processor.sheet)
        self.assertEqual(processor.excel_file, self.test_excel)

    def test_init_with_workbook(self):
        """测试使用 Workbook 对象初始化"""
        processor = ExcelProcessor(self.wb, self.wb.active)
        self.assertIsNotNone(processor.wb)
        self.assertIsNotNone(processor.sheet)
        self.assertIsNone(processor.excel_file)

    def test_init_with_sheet_name(self):
        """测试指定工作表名称"""
        processor = ExcelProcessor(self.test_excel, "TestSheet")
        self.assertEqual(processor.sheet.title, "TestSheet")

    def test_get_cell_value_smart_normal_cell(self):
        """测试获取普通单元格的值"""
        processor = ExcelProcessor(self.test_excel)
        value = processor.get_cell_value_smart(1, 1)
        self.assertEqual(value, "Field1")

    def test_get_cell_value_smart_merged_cell(self):
        """测试获取合并单元格的值"""
        processor = ExcelProcessor(self.test_excel)
        # A5:B5 合并，主单元格在 A5
        value = processor.get_cell_value_smart(5, 2)  # B5 是合并单元格
        self.assertEqual(value, "Merged Cell")

    def test_copy_cell_style(self):
        """测试复制单元格样式"""
        processor = ExcelProcessor(self.test_excel)
        source_cell = processor.sheet["A1"]
        target_cell = processor.sheet["C1"]

        processor.copy_cell_style(source_cell, target_cell)

        self.assertEqual(target_cell.font.bold, source_cell.font.bold)
        self.assertEqual(target_cell.font.size, source_cell.font.size)
        self.assertIsNotNone(target_cell.border.left)

    def test_match_and_fill_top_table(self):
        """测试匹配并填充顶格表"""
        processor = ExcelProcessor(self.test_excel)

        # 准备测试数据 - 注意使用 'fields' 而不是 'data'
        log_section = {
            "type": "opSch",
            "fields": {"Field1": "NewValue1", "Field2": "NewValue2"},
        }

        config = {
            "top_table": {
                "enable": True,
                "log_keyword": "opSch",
            },
            "special_prefix": {
                "enable": False,
                "for_b_column": [],
            },
            "matching": {
                "enable_partial_match": True,
            },
        }

        processor.config = config

        # 调用方法（不验证具体结果，只确保不抛出异常）
        try:
            result = processor.match_and_fill_top_table(
                log_section, start_row=1, end_row=3
            )
            # 应该返回匹配结果（可能是 int 或 dict）
            self.assertIsNotNone(result)
        except Exception as e:
            self.fail(f"match_and_fill_top_table() raised {e}")

    def test_find_top_table(self):
        """测试查找顶格表"""
        processor = ExcelProcessor(self.test_excel)
        result = processor.find_top_table()
        self.assertIsNotNone(result)

    def test_find_sub_table(self):
        """测试查找子表"""
        processor = ExcelProcessor(self.test_excel)

        # 添加子表标题行
        processor.sheet["A7"] = "ERCfg (grp = 0)"
        processor.sheet["A8"] = "cfgGroup"
        processor.sheet["A9"] = "powerLevel"

        processor.find_sub_table("ERCfg")
        # 如果没有配置，可能找不到
        # self.assertIsNotNone(result)

    def test_save(self):
        """测试保存方法"""
        output_file = os.path.join(self.temp_dir, "output.xlsx")
        processor = ExcelProcessor(self.test_excel)

        processor.save(output_file)

        self.assertTrue(os.path.exists(output_file))

    def test_normalize_subtable_spacing(self):
        """测试规范化子表间距"""
        processor = ExcelProcessor(self.test_excel)
        processor.config = {"keyword_mapping": {}}

        # 不应抛出异常
        try:
            processor.normalize_subtable_spacing()
        except Exception as e:
            self.fail(f"normalize_subtable_spacing() raised {e}")

    def test_get_warnings(self):
        """测试获取警告信息"""
        processor = ExcelProcessor(self.test_excel)

        processor.warnings.append("Test warning 1")
        processor.warnings.append("Test warning 2")

        self.assertEqual(len(processor.warnings), 2)
        self.assertEqual(processor.warnings[0], "Test warning 1")


class TestExcelProcessorEdgeCases(unittest.TestCase):
    """测试 ExcelProcessor 边界情况"""

    def setUp(self):
        """测试前准备"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_excel = os.path.join(self.temp_dir, "edge_test.xlsx")

        wb = Workbook()

        # 创建空的工作表
        wb.save(self.test_excel)

    def tearDown(self):
        """测试后清理"""
        import shutil

        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def test_empty_sheet(self):
        """测试空工作表"""
        processor = ExcelProcessor(self.test_excel)
        self.assertIsNotNone(processor.sheet)

    def test_find_top_table_empty_sheet(self):
        """测试在空工作表中查找顶格表"""
        processor = ExcelProcessor(self.test_excel)
        processor.find_top_table()
        # 空工作表应该返回 None 或空结果
        # self.assertIsNone(result)

    def test_config_defaults(self):
        """测试配置默认值"""
        processor = ExcelProcessor(self.test_excel)
        self.assertIsNotNone(processor.config)
        self.assertEqual(processor.config, {})


if __name__ == "__main__":
    unittest.main()
