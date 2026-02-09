#!/usr/bin/env python3
"""
测试 excel_writer 的 multi_top_mode 配置
"""

import os
import sys
import tempfile
import shutil
import unittest

# 添加项目根目录到 Python 路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.workflow import process_log_to_excel


class TestMultiTopMode(unittest.TestCase):
    """测试 multi_top_mode 配置开关"""

    def setUp(self):
        """测试前准备"""
        # 创建临时目录
        self.temp_dir = tempfile.mkdtemp()

        # 创建包含 2 个 opSch 的 trace 文件
        self.trace_content = """thread0	cyc=0x1000 opSch
thread0	cyc=0x1000 |- systemMode    = 1  (0x01)
thread0	cyc=0x1000 |- debugLevel    = 2  (0x02)

thread0	cyc=0x2000 ERCfg (grp = 0)
thread0	cyc=0x2000 |- cfgGroup    = 0  (0x00)
thread0	cyc=0x2000 |- powerLevel  = 15  (0x0F)

thread0	cyc=0x5aa2 opSch
thread0	cyc=0x5aa2 |- systemMode    = 2  (0x02)
thread0	cyc=0x5aa2 |- debugLevel    = 3  (0x03)

thread0	cyc=0x6000 TXCfg (grp = 0)
thread0	cyc=0x6000 |- cfgGroup    = 0  (0x00)
thread0	cyc=0x6000 |- txPower     = 30  (0x1E)
"""
        self.trace_file = os.path.join(self.temp_dir, "test_trace.txt")
        with open(self.trace_file, "w", encoding="utf-8") as f:
            f.write(self.trace_content)

        # 使用现有模板
        self.template_file = "templates/template_a_column.xlsx"
        if not os.path.exists(self.template_file):
            self.template_file = "examples/templates/template_a_column.xlsx"

    def tearDown(self):
        """测试后清理"""
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def test_multi_sheets_mode(self):
        """测试分页签模式（默认）"""
        print("\n" + "=" * 60)
        print("测试 1: multi_sheets 模式（分页签）")
        print("=" * 60)

        output_file = os.path.join(self.temp_dir, "output_multi_sheets.xlsx")

        # 执行 workflow（默认配置 multi_top_mode: 'multi_sheets'）
        result_file = process_log_to_excel(
            excel_file=self.template_file,
            trace_file=self.trace_file,
            output_file=output_file,
        )

        # 验证结果
        self.assertIsNotNone(result_file)
        self.assertTrue(os.path.exists(result_file))

        # 验证是否创建了多个工作表
        from openpyxl import load_workbook

        wb = load_workbook(result_file)
        sheet_names = wb.sheetnames
        print("✓ 测试通过！")
        print(f"  输出文件: {result_file}")
        print(f"  工作表数量: {len(sheet_names)}")
        print(f"  工作表名称: {sheet_names}")

        # 应该有 2 个工作表（2 个 opSch）
        self.assertGreaterEqual(len(sheet_names), 2, "应该创建至少 2 个工作表")

    def test_single_sheet_mode(self):
        """测试单页签模式（分表格）"""
        print("\n" + "=" * 60)
        print("测试 2: single_sheet 模式（分表格）")
        print("=" * 60)

        # 临时修改配置
        config_file = "src/plugins/excel_writer/config.yaml"
        import yaml

        # 读取原始配置
        with open(config_file, "r", encoding="utf-8") as f:
            original_config = f.read()

        try:
            # 修改配置为 single_sheet 模式
            with open(config_file, "r", encoding="utf-8") as f:
                config = yaml.safe_load(f)

            config["top_table"]["multi_top_mode"] = "single_sheet"

            with open(config_file, "w", encoding="utf-8") as f:
                yaml.dump(config, f, allow_unicode=True)

            # 重新加载插件（清除缓存）
            import importlib
            import src.plugins.excel_writer.plugin

            importlib.reload(src.plugins.excel_writer.plugin)

            output_file = os.path.join(self.temp_dir, "output_single_sheet.xlsx")

            # 执行 workflow
            result_file = process_log_to_excel(
                excel_file=self.template_file,
                trace_file=self.trace_file,
                output_file=output_file,
            )

            # 验证结果
            self.assertIsNotNone(result_file)
            self.assertTrue(os.path.exists(result_file))

            # 验证只有 1 个工作表
            from openpyxl import load_workbook

            wb = load_workbook(result_file)
            sheet_names = wb.sheetnames
            print("✓ 测试通过！")
            print(f"  输出文件: {result_file}")
            print(f"  工作表数量: {len(sheet_names)}")
            print(f"  工作表名称: {sheet_names}")

            # 应该只有 1 个工作表（所有数据在同一个表）
            self.assertEqual(len(sheet_names), 1, "应该只有 1 个工作表")

        finally:
            # 恢复原始配置
            with open(config_file, "w", encoding="utf-8") as f:
                f.write(original_config)

            # 重新加载插件
            importlib.reload(src.plugins.excel_writer.plugin)

    def test_config_validation(self):
        """测试配置值的验证"""
        print("\n" + "=" * 60)
        print("测试 3: 配置读取")
        print("=" * 60)

        from src.plugins.excel_writer import ExcelWriterPlugin

        plugin = ExcelWriterPlugin()
        config = plugin.config

        multi_top_mode = config.get("top_table", {}).get(
            "multi_top_mode", "multi_sheets"
        )
        print("✓ 当前配置:")
        print(f"  multi_top_mode: {multi_top_mode}")

        self.assertIn(
            multi_top_mode,
            ["multi_sheets", "single_sheet"],
            "配置值应该是 'multi_sheets' 或 'single_sheet'",
        )


def run_tests():
    """运行所有测试"""
    print("\n")
    print("=" * 60)
    print("multi_top_mode 配置测试")
    print("=" * 60)

    # 创建测试套件
    suite = unittest.TestLoader().loadTestsFromTestCase(TestMultiTopMode)

    # 运行测试
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    # 打印总结
    print("\n" + "=" * 60)
    print("测试总结")
    print("=" * 60)
    print(f"运行测试数: {result.testsRun}")
    print(f"成功: {result.testsRun - len(result.failures) - len(result.errors)}")
    print(f"失败: {len(result.failures)}")
    print(f"错误: {len(result.errors)}")
    print("=" * 60)

    return result.wasSuccessful()


if __name__ == "__main__":
    success = run_tests()
    sys.exit(0 if success else 1)
