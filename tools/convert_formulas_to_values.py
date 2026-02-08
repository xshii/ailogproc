#!/usr/bin/env python3
"""
将Excel模板中的公式转换为静态值

用法：
    python tools/convert_formulas_to_values.py template.xlsx

这个脚本会：
1. 读取Excel文件
2. 找到所有包含公式的单元格
3. 计算公式的值
4. 用计算后的值替换公式
5. 保存为新文件（原文件备份为 .bak）
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


def convert_formulas_to_values(excel_file, sheet_name=None):
    """将Excel中的公式转换为值"""
    print(f"正在处理: {excel_file}")

    # 加载工作簿
    wb = load_workbook(excel_file, data_only=False)  # data_only=False 读取公式

    # 选择工作表
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"错误: 找不到工作表 '{sheet_name}'")
            print(f"可用工作表: {wb.sheetnames}")
            return
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    print(f"工作表: {sheet.title}")

    # 查找并转换公式
    formula_count = 0
    converted_cells = []

    for row in sheet.iter_rows():
        for cell in row:
            # 跳过合并单元格
            if isinstance(cell, MergedCell):
                continue

            # 检查是否是公式
            if cell.data_type == 'f':
                formula = cell.value

                # 尝试计算值（openpyxl不会自动计算，所以值可能为None）
                # 这里我们需要用data_only=True重新读取来获取计算值
                formula_count += 1
                converted_cells.append((cell.coordinate, formula))

    if formula_count == 0:
        print("✓ 没有找到公式，无需转换")
        return

    print(f"\n找到 {formula_count} 个公式单元格:")
    for coord, formula in converted_cells[:10]:  # 只显示前10个
        print(f"  {coord}: {formula}")
    if len(converted_cells) > 10:
        print(f"  ... 还有 {len(converted_cells) - 10} 个")

    # 重新加载工作簿（data_only=True 获取计算值）
    print("\n正在计算公式值...")
    wb_values = load_workbook(excel_file, data_only=True)

    if sheet_name:
        sheet_values = wb_values[sheet_name]
    else:
        sheet_values = wb_values.active

    # 替换公式为值
    converted = 0
    for coord, formula in converted_cells:
        cell = sheet[coord]
        value_cell = sheet_values[coord]
        value = value_cell.value

        # 替换公式为值
        cell.value = value
        converted += 1

    print(f"✓ 已转换 {converted} 个公式为值")

    # 备份原文件
    excel_path = Path(excel_file)
    backup_path = excel_path.with_suffix(excel_path.suffix + '.bak')

    if backup_path.exists():
        print(f"⚠️  备份文件已存在: {backup_path}")
    else:
        excel_path.rename(backup_path)
        print(f"✓ 原文件备份为: {backup_path}")

    # 保存新文件
    wb.save(excel_file)
    print(f"✓ 已保存: {excel_file}")
    print("\n完成！")


def main():
    if len(sys.argv) < 2:
        print("用法: python tools/convert_formulas_to_values.py <Excel文件> [工作表名]")
        print()
        print("示例:")
        print("  python tools/convert_formulas_to_values.py template.xlsx")
        print("  python tools/convert_formulas_to_values.py template.xlsx Sheet1")
        sys.exit(1)

    excel_file = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None

    if not Path(excel_file).exists():
        print(f"错误: 文件不存在: {excel_file}")
        sys.exit(1)

    convert_formulas_to_values(excel_file, sheet_name)


if __name__ == "__main__":
    main()
