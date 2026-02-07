"""
业务流程模块 - 主要处理逻辑
"""
import os
import re
from openpyxl import load_workbook

from config.default_config import (
    KEYWORD_MAPPING,
    ENABLE_TOP_TABLE,
    TOP_TABLE_LOG_KEYWORD,
    ENABLE_AUTO_FILENAME,
    FILENAME_FIELDS,
    FILENAME_VALUE_MAPPING,
    FILENAME_DEFAULT_VALUE,
    SPECIAL_PREFIX_FOR_B_COLUMN,
    SPECIAL_PREFIX_MERGE_ROWS,
    ENABLE_PARTIAL_MATCH,
)
from src.log_parser import LogParser
from src.excel_processor import ExcelProcessor
from src.utils import get_target_column



def process_log_to_excel(excel_file, log_file, output_file=None, sheet_name=None):
    """
    主处理流程：解析日志并填充到Excel
    支持单个TopConfig或多个TopConfig（多工作表）
    """
    print("=" * 60)
    print("开始处理...")
    print("=" * 60)
    
    # 1. 解析日志
    print("\n[1/5] 解析日志文件...")
    parser = LogParser(log_file)
    sections = parser.parse()
    print(f"✓ 解析完成，找到 {len(sections)} 个配置块")
    
    # 检查是否有多个TopConfig
    if ENABLE_TOP_TABLE:
        groups = parser.group_by_top_config(TOP_TABLE_LOG_KEYWORD)
        num_top_configs = len(groups)
        
        if num_top_configs > 1:
            print(f"\n📋 检测到 {num_top_configs} 个 {TOP_TABLE_LOG_KEYWORD} 配置块")
            print(f"   将创建 {num_top_configs} 个工作表页签")
            return process_multi_sheets(excel_file, log_file, output_file, sheet_name, parser, groups)
        elif num_top_configs == 1:
            print(f"\n📋 检测到 1 个 {TOP_TABLE_LOG_KEYWORD} 配置块")
            print(f"   将使用单工作表模式")
        else:
            print(f"\n⚠️  未检测到 {TOP_TABLE_LOG_KEYWORD} 配置块")
    
    # 单工作表模式（原有逻辑）
    return process_single_sheet(excel_file, log_file, output_file, sheet_name, parser, sections)


def process_single_sheet(excel_file, log_file, output_file, sheet_name, parser, sections):
    """
    单工作表处理模式（原有逻辑）
    """
    # 2. 加载Excel
    print("\n[2/5] 加载Excel文件...")
    processor = ExcelProcessor(excel_file, sheet_name)
    print(f"✓ 加载完成，使用工作表: {processor.sheet.title}")
    
    # 3. 处理顶格表格
    print("\n[3/5] 处理顶格表格...")
    if ENABLE_TOP_TABLE:
        # 使用日志中的关键字获取配置块
        top_section = parser.get_top_section(TOP_TABLE_LOG_KEYWORD)
        if top_section:
            # Excel中直接从配置的起始行开始查找
            start_row, end_row = processor.find_top_table()
            if start_row:
                matched = processor.match_and_fill_top_table(top_section, start_row, end_row)
                print(f"✓ 顶格表格处理完成，匹配了 {len(matched)} 个字段")
            else:
                print(f"✗ Excel中未找到顶格表格（起始行: {TOP_TABLE_START_ROW}）")
        else:
            print(f"✗ 日志中未找到顶格配置块: {TOP_TABLE_LOG_KEYWORD}")
    else:
        print(f"✗ 顶格表格处理已禁用（ENABLE_TOP_TABLE = False）")
    
    # 4. 处理子片段（按日志顺序）
    print("\n[4/5] 处理子片段...")
    
    # 记录每个Excel关键词的原始位置和使用次数
    keyword_info = {}  # {excel_keyword: {'orig_start', 'orig_end', 'count', 'used'}}
    
    # 先扫描所有原始子表的位置
    for excel_keyword in KEYWORD_MAPPING.keys():
        start_row, end_row = processor.find_sub_table(excel_keyword)
        if start_row:
            keyword_info[excel_keyword] = {
                'orig_start': start_row,
                'orig_end': end_row,
                'count': 0,
                'used': False
            }
    
    # 找到当前工作表的最后一行（用于追加）
    global_last_row = processor.sheet.max_row
    
    # 按日志顺序处理每个配置块
    for section in sections:
        section_name = section['name']
        
        # 跳过顶格配置块
        if section_name == TOP_TABLE_LOG_KEYWORD:
            continue
        
        # 查找匹配的Excel关键词
        matched_excel_keyword = None
        for excel_keyword, log_pattern in KEYWORD_MAPPING.items():
            if re.match(log_pattern, section_name):
                matched_excel_keyword = excel_keyword
                break
        
        if not matched_excel_keyword:
            print(f"  - {section_name}: 未找到对应的Excel关键词")
            continue
        
        if matched_excel_keyword not in keyword_info:
            print(f"  - {section_name} → {matched_excel_keyword}: Excel中未找到子表")
            continue
        
        info = keyword_info[matched_excel_keyword]
        info['count'] += 1
        info['used'] = True
        
        # 所有子表都复制到末尾（保持日志顺序）
        new_end = processor.copy_sub_table(
            info['orig_start'], info['orig_end'], global_last_row
        )
        
        # 新表格的起始行（关键字行）
        new_table_start = global_last_row + 2  # 空行后的第一行
        
        matched = processor.match_and_fill_sub_table(
            section, new_table_start, new_end
        )
        
        # 在关键字行的目标列填充完整的日志关键字名称
        target_col = get_target_column()
        keyword_row = new_table_start  # 关键字行就是表格起始行
        processor.sheet.cell(keyword_row, target_col, value=section_name)
        
        print(f"  - {section_name} → {matched_excel_keyword} (#{info['count']}): 匹配 {len(matched)} 个字段")
        
        # 更新全局最后一行
        global_last_row = new_end

    
    # 删除未使用的子表（从后往前删除，避免行号变化）
    print(f"\n  清理未使用的子表...")
    unused_tables = []
    for excel_keyword, info in keyword_info.items():
        if not info['used']:
            unused_tables.append((excel_keyword, info))
    
    if unused_tables:
        # 按起始行倒序排列（从后往前删除）
        unused_tables.sort(key=lambda x: x[1]['orig_start'], reverse=True)
        
        for excel_keyword, info in unused_tables:
            start_row = info['orig_start']
            end_row = info['orig_end']
            
            # 只删除表格本身，不管前后空行（后续统一规范化）
            num_rows = end_row - start_row + 1
            processor.sheet.delete_rows(start_row, num_rows)
            print(f"    ✓ 已删除 {excel_keyword} (第{start_row}-{end_row}行，共{num_rows}行)")
        
        # 删除完成后，规范化所有子表之间的空行
        print(f"\n  规范化子表间距...")
        processor.normalize_subtable_spacing()
    else:
        print(f"    ✓ 所有子表都已使用，无需清理")
    
    # 3. 处理顶格表格
    print("\n[3/5] 处理顶格表格...")
    if ENABLE_TOP_TABLE:
        # 使用日志中的关键字获取配置块
        top_section = parser.get_top_section(TOP_TABLE_LOG_KEYWORD)
        if top_section:
            # Excel中直接从配置的起始行开始查找
            start_row, end_row = processor.find_top_table()
            if start_row:
                matched = processor.match_and_fill_top_table(top_section, start_row, end_row)
                print(f"✓ 顶格表格处理完成，匹配了 {len(matched)} 个字段")
            else:
                print(f"✗ Excel中未找到顶格表格（起始行: {TOP_TABLE_START_ROW}）")
        else:
            print(f"✗ 日志中未找到顶格配置块: {TOP_TABLE_LOG_KEYWORD}")
    else:
        print(f"✗ 顶格表格处理已禁用（ENABLE_TOP_TABLE = False）")
    
    # 4. 处理子片段（按日志顺序）
    print("\n[4/5] 处理子片段...")
    
    # 记录每个Excel关键词的原始位置和使用次数
    keyword_info = {}  # {excel_keyword: {'orig_start', 'orig_end', 'count', 'used'}}
    
    # 先扫描所有原始子表的位置
    for excel_keyword in KEYWORD_MAPPING.keys():
        start_row, end_row = processor.find_sub_table(excel_keyword)
        if start_row:
            keyword_info[excel_keyword] = {
                'orig_start': start_row,
                'orig_end': end_row,
                'count': 0,
                'used': False
            }
    
    # 找到当前工作表的最后一行（用于追加）
    global_last_row = processor.sheet.max_row
    
    # 按日志顺序处理每个配置块
    for section in sections:
        section_name = section['name']
        
        # 跳过顶格配置块
        if section_name == TOP_TABLE_LOG_KEYWORD:
            continue
        
        # 查找匹配的Excel关键词
        matched_excel_keyword = None
        for excel_keyword, log_pattern in KEYWORD_MAPPING.items():
            if re.match(log_pattern, section_name):
                matched_excel_keyword = excel_keyword
                break
        
        if not matched_excel_keyword:
            print(f"  - {section_name}: 未找到对应的Excel关键词")
            continue
        
        if matched_excel_keyword not in keyword_info:
            print(f"  - {section_name} → {matched_excel_keyword}: Excel中未找到子表")
            continue
        
            info = keyword_info[matched_excel_keyword]
            info['count'] += 1
            info['used'] = True
            
            # 所有子表都复制到末尾（保持日志顺序）
            new_end = processor.copy_sub_table(
                info['orig_start'], info['orig_end'], global_last_row
            )
            
            # 新表格的起始行（关键字行）
            new_table_start = global_last_row + 2  # 空行后的第一行
            
            matched = processor.match_and_fill_sub_table(
                section, new_table_start, new_end
            )
            
            # 在关键字行的目标列填充完整的日志关键字名称
            target_col = get_target_column()
            processor.sheet.cell(new_table_start, target_col, value=section_name)
            
            print(f"     - {section_name} → {matched_excel_keyword} (#{info['count']}): 匹配 {len(matched)} 个字段")
            
            # 更新全局最后一行
            global_last_row = new_end

    
    # 删除未使用的子表（从后往前删除，避免行号变化）
    print(f"\n  清理未使用的子表...")
    unused_tables = []
    for excel_keyword, info in keyword_info.items():
        if not info['used']:
            unused_tables.append((excel_keyword, info))
    
    if unused_tables:
        # 按起始行倒序排列（从后往前删除）
        unused_tables.sort(key=lambda x: x[1]['orig_start'], reverse=True)
        
        for excel_keyword, info in unused_tables:
            start_row = info['orig_start']
            end_row = info['orig_end']
            
            # 只删除表格本身，不管前后空行（后续统一规范化）
            num_rows = end_row - start_row + 1
            processor.sheet.delete_rows(start_row, num_rows)
            print(f"    ✓ 已删除 {excel_keyword} (第{start_row}-{end_row}行，共{num_rows}行)")
        
        # 删除完成后，规范化所有子表之间的空行
        print(f"\n  规范化子表间距...")
        processor.normalize_subtable_spacing()
    else:
        print(f"    ✓ 所有子表都已使用，无需清理")
    
    # 5. 保存文件
    print("\n[5/5] 保存文件...")
    if not output_file:
        import os
        base_name = os.path.splitext(excel_file)[0]
        output_file = f"{base_name}_processed.xlsx"
    
    # 自动生成文件名后缀
    if ENABLE_AUTO_FILENAME and ENABLE_TOP_TABLE:
        output_file = generate_auto_filename(output_file, processor)
    
    processor.save(output_file)
    
    # 显示告警
    if processor.warnings:
        print("\n" + "=" * 60)
        print("⚠️  告警信息:")
        print("=" * 60)
        for warning in processor.warnings:
            print(warning)
    
    print("\n" + "=" * 60)
    print("✅ 处理完成!")
    print("=" * 60)
    
    return output_file


# ==================== 命令行入口 ====================

def generate_auto_filename(original_filename, processor):
    """根据顶格表格的字段值自动生成文件名后缀
    
    Args:
        original_filename: 原始输出文件名
        processor: ExcelProcessor实例
        
    Returns:
        新的文件名（带字段值后缀）
    """
    import os
    
    print("\n  生成自动文件名...")
    
    # 如果没有配置字段，返回原文件名
    if not FILENAME_FIELDS:
        print("  ⚠️  未配置FILENAME_FIELDS，使用原文件名")
        return original_filename
    
    # 获取顶格表格范围
    if not ENABLE_TOP_TABLE:
        print("  ⚠️  顶格表格未启用，使用原文件名")
        return original_filename
    
    top_start, top_end = processor.find_top_table()
    if not top_start:
        print("  ⚠️  未找到顶格表格，使用原文件名")
        return original_filename
    
    # 提取字段值
    target_col = get_target_column()
    field_values = []
    
    for field_name in FILENAME_FIELDS:
        # 在顶格表格中查找字段名（使用与match_and_fill_top_table相同的逻辑）
        found = False
        field_value = None
        field_name_lower = field_name.lower()
        
        for row in range(top_start, top_end + 1):
            # 检查A列
            a_col_value = processor.get_cell_value_smart(row, 1)  # A列
            
            if not a_col_value:
                continue
            
            a_col_str = str(a_col_value).strip()
            
            # 检查是否是特殊前缀
            is_special_prefix = any(
                a_col_str.startswith(prefix) 
                for prefix in SPECIAL_PREFIX_FOR_B_COLUMN
            )
            
            if is_special_prefix:
                # 特殊前缀：匹配B列
                b_col_value = processor.get_cell_value_smart(row, 2)
                if b_col_value:
                    b_col_str = str(b_col_value).strip()
                    b_col_str_lower = b_col_str.lower()
                    
                    if b_col_str_lower == field_name_lower:
                        # 找到字段，读取目标列的值
                        field_value = processor.get_cell_value_smart(row, target_col)
                        found = True
                        break
            else:
                # 普通：匹配A列
                a_col_str_lower = a_col_str.lower()
                
                if a_col_str_lower == field_name_lower:
                    # 找到字段，读取目标列的值
                    field_value = processor.get_cell_value_smart(row, target_col)
                    found = True
                    break
        
        if not found or field_value is None:
            print(f"  ⚠️  字段 '{field_name}' 未找到或无值，跳过")
            continue
        
        # 转换字段值为字符串
        field_value_str = str(field_value).strip()
        
        # 应用映射
        if field_name in FILENAME_VALUE_MAPPING:
            field_mapping = FILENAME_VALUE_MAPPING[field_name]
            if field_value_str in field_mapping:
                mapped_value = field_mapping[field_value_str]
                field_values.append(mapped_value)
                print(f"  ✓ {field_name}: {field_value_str} → {mapped_value}")
            else:
                # 未找到映射，使用默认值或原值
                if FILENAME_DEFAULT_VALUE:
                    field_values.append(FILENAME_DEFAULT_VALUE)
                    print(f"  ⚠️  {field_name}: {field_value_str} → {FILENAME_DEFAULT_VALUE} (未映射)")
                else:
                    field_values.append(field_value_str)
                    print(f"  ⚠️  {field_name}: {field_value_str} (未映射，使用原值)")
        else:
            # 该字段没有映射表，直接使用原值
            field_values.append(field_value_str)
            print(f"  ℹ️  {field_name}: {field_value_str} (无映射配置)")
    
    # 如果没有提取到任何字段值，返回原文件名
    if not field_values:
        print("  ⚠️  未提取到任何字段值，使用原文件名")
        return original_filename
    
    # 生成文件名后缀
    suffix = '_' + '_'.join(field_values)
    
    # 分解原文件名
    dir_name = os.path.dirname(original_filename)
    base_name = os.path.basename(original_filename)
    name_without_ext, ext = os.path.splitext(base_name)
    
    # 生成新文件名
    new_base_name = name_without_ext + suffix + ext
    new_filename = os.path.join(dir_name, new_base_name) if dir_name else new_base_name
    
    print(f"  ✓ 新文件名: {new_base_name}")
    
    return new_filename





def generate_sheet_name(top_section, index):
    """根据TopConfig生成工作表名称"""
    # 尝试从TopConfig字段生成名称
    if FILENAME_FIELDS and ENABLE_AUTO_FILENAME:
        name_parts = []
        for field_name in FILENAME_FIELDS[:2]:  # 只用前2个字段
            if field_name in top_section['fields']:
                value = top_section['fields'][field_name]
                # 应用映射
                if field_name in FILENAME_VALUE_MAPPING:
                    field_mapping = FILENAME_VALUE_MAPPING[field_name]
                    value_str = str(value).strip()
                    if value_str in field_mapping:
                        value = field_mapping[value_str]
                name_parts.append(str(value))
        
        if name_parts:
            sheet_name = f"Config_{index}_{'_'.join(name_parts)}"
            # Excel工作表名称限制：31字符
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:28] + "..."
            return sheet_name
    
    # 默认使用序号
    return f"Config_{index}"


def process_multi_sheets(excel_file, log_file, output_file, sheet_name, parser, groups):
    """多工作表处理模式：为每个TopConfig创建独立页签"""
    from openpyxl import load_workbook
    import os
    
    print("\n[2/5] 加载Excel模板...")
    workbook = load_workbook(excel_file)
    
    # 获取模板工作表
    if sheet_name and sheet_name in workbook.sheetnames:
        template_sheet = workbook[sheet_name]
    else:
        template_sheet = workbook.active
    
    template_sheet_name = template_sheet.title
    print(f"✓ 使用模板工作表: {template_sheet_name}")
    
    # 为每个TopConfig组创建工作表
    print(f"\n[3/5] 创建 {len(groups)} 个工作表...")
    
    created_sheets = []
    
    for idx, group in enumerate(groups, 1):
        top_section = group['top']
        sub_sections = group['subs']
        
        # 生成工作表名称
        sheet_title = generate_sheet_name(top_section, idx)
        print(f"\n  [{idx}/{len(groups)}] 创建工作表: {sheet_title}")
        
        # 复制模板工作表
        new_sheet = workbook.copy_worksheet(template_sheet)
        new_sheet.title = sheet_title
        
        # 创建processor处理这个工作表
        processor = ExcelProcessor(excel_file, template_sheet_name)
        processor.workbook = workbook
        processor.sheet = new_sheet
        
        # 处理顶格表格
        if ENABLE_TOP_TABLE:
            start_row, end_row = processor.find_top_table()
            if start_row:
                matched = processor.match_and_fill_top_table(top_section, start_row, end_row)
                print(f"     ✓ 顶格表格: 匹配 {len(matched)} 个字段")
        
        # 处理子表
        keyword_info = {}
        for excel_keyword in KEYWORD_MAPPING.keys():
            start_row, end_row = processor.find_sub_table(excel_keyword)
            if start_row:
                keyword_info[excel_keyword] = {
                    'orig_start': start_row,
                    'orig_end': end_row,
                    'count': 0,
                    'used': False
                }
        
        global_last_row = processor.sheet.max_row
        
        # 处理这个组的子配置块
        for section in sub_sections:
            section_name = section['name']
            
            # 查找匹配的Excel关键词
            matched_excel_keyword = None
            for excel_keyword, log_pattern in KEYWORD_MAPPING.items():
                if re.match(log_pattern, section_name):
                    matched_excel_keyword = excel_keyword
                    break
            
            if not matched_excel_keyword or matched_excel_keyword not in keyword_info:
                continue
            
            info = keyword_info[matched_excel_keyword]
            info['count'] += 1
            info['used'] = True
            
            if info['count'] == 1:
                matched = processor.match_and_fill_sub_table(
                    section, info['orig_start'], info['orig_end']
                )
                target_col = get_target_column()
                processor.sheet.cell(info['orig_start'], target_col, value=section_name)
            else:
                new_end = processor.copy_sub_table(
                    info['orig_start'], info['orig_end'], global_last_row
                )
                new_table_start = global_last_row + 2
                matched = processor.match_and_fill_sub_table(
                    section, new_table_start, new_end
                )
                target_col = get_target_column()
                processor.sheet.cell(new_table_start, target_col, value=section_name)
                global_last_row = new_end
        
        # 删除未使用的子表
        unused_tables = [(k, v) for k, v in keyword_info.items() if not v['used']]
        if unused_tables:
            unused_tables.sort(key=lambda x: x[1]['orig_start'], reverse=True)
            for excel_keyword, info in unused_tables:
                num_rows = info['orig_end'] - info['orig_start'] + 1
                processor.sheet.delete_rows(info['orig_start'], num_rows)
            processor.normalize_subtable_spacing()
        
        sub_count = len([s for s in sub_sections if any(re.match(p, s['name']) for p in KEYWORD_MAPPING.values())])
        print(f"     ✓ 处理了 {sub_count} 个子表")
        
        created_sheets.append((sheet_title, processor))
    
    # 删除原模板工作表
    if template_sheet_name in workbook.sheetnames:
        del workbook[template_sheet_name]
        print(f"\n  ✓ 已删除模板工作表: {template_sheet_name}")
    
    # 保存文件
    print(f"\n[4/5] 保存文件...")
    if not output_file:
        base_name = os.path.splitext(excel_file)[0]
        output_file = f"{base_name}_multi_sheets.xlsx"
    
    # 自动生成文件名（基于第一个top表的字段值）
    if ENABLE_AUTO_FILENAME and ENABLE_TOP_TABLE and created_sheets:
        first_processor = created_sheets[0][1]
        output_file = generate_auto_filename(output_file, first_processor)
    
    workbook.save(output_file)
    
    print(f"\n{'='*60}")
    print("✅ 保存成功!")
    print(f"{'='*60}")
    print(f"文件位置: {output_file}")
    print(f"工作表数量: {len(created_sheets)}")
    for sheet_name, _ in created_sheets:
        print(f"  - {sheet_name}")
    print(f"{'='*60}")
    
    print("\n" + "=" * 60)
    print("✅ 处理完成!")
    print("=" * 60)


