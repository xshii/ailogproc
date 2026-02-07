"""
Excel处理器模块
"""
import re
from copy import copy
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from config.default_config import (
    KEYWORD_MAPPING,
    TOP_TABLE_START_ROW,
    ENABLE_PARTIAL_MATCH,
    SHOW_UNMATCHED_WARNINGS,
    SPECIAL_PREFIX_FOR_B_COLUMN,
    SPECIAL_PREFIX_MERGE_ROWS,
)
from src.utils import get_target_column


class ExcelProcessor:
    """处理Excel表格的读取、匹配和填充"""
    
    def __init__(self, excel_file, sheet_name=None):
        self.excel_file = excel_file
        self.wb = load_workbook(excel_file)
        self.sheet = self.wb[sheet_name] if sheet_name else self.wb.active
        self.warnings = []  # 存储告警信息
    
    def get_cell_value_smart(self, row, col):
        """智能获取单元格值，处理合并单元格
        
        如果是MergedCell，找到合并区域的主单元格并返回其值
        
        Args:
            row: 行号（从1开始）
            col: 列号（从1开始）
            
        Returns:
            单元格的值，如果是MergedCell则返回主单元格的值
        """
        from openpyxl.cell.cell import MergedCell
        
        cell = self.sheet.cell(row, col)
        
        # 如果是普通单元格，直接返回值
        if not isinstance(cell, MergedCell):
            return cell.value
        
        # 如果是MergedCell，找到对应的合并区域
        for merged_range in self.sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 获取合并区域的主单元格（左上角）
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                master_cell = self.sheet.cell(min_row, min_col)
                return master_cell.value
        
        # 如果没有找到合并区域（不应该发生），返回None
        return None
        
    def copy_cell_style(self, source_cell, target_cell):
        """复制单元格样式"""
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    
    def normalize_subtable_spacing(self):
        """规范化所有子表之间的空行，确保每两个子表之间有且只有一个空行"""
        # 找到所有子表的位置
        subtable_positions = []
        
        row = 1
        while row <= self.sheet.max_row:
            cell_a = self.sheet.cell(row, 1).value
            # 检查是否是子表关键字行
            if cell_a and any(keyword in str(cell_a) for keyword in KEYWORD_MAPPING.keys()):
                start_row = row
                # 找到子表的结束行（最后一个有数据的行）
                end_row = start_row
                for check_row in range(start_row + 1, min(start_row + 20, self.sheet.max_row + 1)):
                    cell_check_a = self.sheet.cell(check_row, 1).value
                    
                    # 如果遇到另一个子表关键字，当前表结束（不修改end_row，保持为最后一个有数据的行）
                    if cell_check_a and any(keyword in str(cell_check_a) for keyword in KEYWORD_MAPPING.keys()):
                        break
                    
                    cell_b = self.sheet.cell(check_row, 2).value
                    cell_c = self.sheet.cell(check_row, 3).value
                    cell_d = self.sheet.cell(check_row, 4).value
                    
                    # 如果B/C/D列有任何内容，说明还在表格内
                    if cell_b or cell_c or cell_d:
                        end_row = check_row
                
                subtable_positions.append((start_row, end_row))
                # 跳过这个表格，从结束行后面继续
                row = end_row + 1
            else:
                row += 1
        
        if len(subtable_positions) <= 1:
            return  # 只有一个或没有子表，无需处理
        
        # 从后往前处理，确保每两个子表之间有且只有一个空行
        for i in range(len(subtable_positions) - 1, 0, -1):
            current_start, current_end = subtable_positions[i]
            prev_start, prev_end = subtable_positions[i - 1]
            
            # 计算两个表之间的空行数
            gap = current_start - prev_end - 1
            
            if gap < 1:
                # 没有空行，插入一个
                self.sheet.insert_rows(prev_end + 1, 1)
                print(f"    ✓ 在第{prev_end}行后插入空行")
                # 更新后续表格位置
                for j in range(i, len(subtable_positions)):
                    old_start, old_end = subtable_positions[j]
                    subtable_positions[j] = (old_start + 1, old_end + 1)
            elif gap > 1:
                # 多个空行，删除多余的（保留1个）
                rows_to_delete = gap - 1
                delete_start = prev_end + 2  # 跳过第一个空行
                self.sheet.delete_rows(delete_start, rows_to_delete)
                print(f"    ✓ 删除多余空行（第{delete_start}行开始，共{rows_to_delete}行）")
                # 更新后续表格位置
                for j in range(i, len(subtable_positions)):
                    old_start, old_end = subtable_positions[j]
                    subtable_positions[j] = (old_start - rows_to_delete, old_end - rows_to_delete)
    
    def find_top_table(self):
        """查找顶格表格的起始和结束行（顶格表格从第1行开始，无需关键字）"""
        start_row = TOP_TABLE_START_ROW
        
        # 查找结束行（遇到空行或子片段关键字）
        end_row = start_row
        for row in range(start_row, self.sheet.max_row + 1):
            # 检查前3列是否都为空（空行）
            if all(self.sheet.cell(row, col).value is None for col in range(1, 4)):
                end_row = row - 1
                break
            
            # 检查是否遇到子片段关键字（A列）
            first_cell = self.sheet.cell(row, 1).value
            if first_cell and str(first_cell).strip() in KEYWORD_MAPPING:
                end_row = row - 1
                break
            
            end_row = row
        
        # 确保至少有一行
        if end_row < start_row:
            return None, None
            
        return start_row, end_row
    
    def find_sub_table(self, keyword):
        """查找子片段表格的起始和结束行"""
        start_row = None
        
        # 查找关键字所在行
        for row in range(1, self.sheet.max_row + 1):
            cell_value = self.sheet.cell(row, 1).value
            if cell_value and str(cell_value).strip() == keyword:
                start_row = row
                break
        
        if not start_row:
            return None, None
        
        # 子表的结束行（遇到空行或新的关键字）
        end_row = start_row
        for row in range(start_row + 1, self.sheet.max_row + 1):
            # 检查A列和B列是否都为空
            if self.sheet.cell(row, 1).value is None and self.sheet.cell(row, 2).value is None:
                break
            # 检查是否遇到新的子片段关键字
            first_cell = self.sheet.cell(row, 1).value
            if first_cell and str(first_cell).strip() in KEYWORD_MAPPING:
                break
            end_row = row
        
        return start_row, end_row
    
    def match_and_fill_top_table(self, log_section, start_row, end_row):
        """在顶格表格中匹配并填充数据
        
        匹配规则：
        1. 默认匹配A列
        2. 如果A列以特殊前缀开头（如'Spec'）→ 改为匹配B列
        3. 特殊前缀匹配时，填写合并2行
        """
        target_col = get_target_column()
        matched_count = {}
        unmatched_fields = []
        special_prefix_no_match = []  # 记录特殊前缀但B列匹配不上的情况
        
        for field_name, field_value in log_section['fields'].items():
            match_info = None  # (row, is_special_prefix)
            field_name_lower = field_name.lower()
            
            # 遍历每一行，决定匹配策略
            for row in range(start_row, end_row + 1):
                a_col_value = self.get_cell_value_smart(row, 1)  # A列
                
                if not a_col_value:
                    continue
                
                a_col_str = str(a_col_value).strip()
                
                # 检查A列是否以特殊前缀开头
                is_special_prefix = any(
                    a_col_str.startswith(prefix) 
                    for prefix in SPECIAL_PREFIX_FOR_B_COLUMN
                )
                
                if is_special_prefix:
                    # 特殊前缀：匹配B列
                    b_col_value = self.get_cell_value_smart(row, 2)  # B列
                    
                    if b_col_value:
                        b_col_str = str(b_col_value).strip()
                        b_col_str_lower = b_col_str.lower()
                        
                        # 精确匹配
                        if b_col_str_lower == field_name_lower:
                            match_info = (row, True)  # True表示特殊前缀
                            break
                        # 部分匹配
                        elif ENABLE_PARTIAL_MATCH and (
                            field_name_lower in b_col_str_lower or 
                            b_col_str_lower in field_name_lower
                        ):
                            match_info = (row, True)
                            break
                    else:
                        # A列有特殊前缀，但B列为空
                        if row not in [info[0] for info in special_prefix_no_match]:
                            special_prefix_no_match.append((row, a_col_str, field_name))
                else:
                    # 普通情况：匹配A列
                    a_col_str_lower = a_col_str.lower()
                    
                    # 精确匹配
                    if a_col_str_lower == field_name_lower:
                        match_info = (row, False)  # False表示普通匹配
                        break
                    # 部分匹配
                    elif ENABLE_PARTIAL_MATCH and (
                        field_name_lower in a_col_str_lower or 
                        a_col_str_lower in field_name_lower
                    ):
                        match_info = (row, False)
                        break
            
            # 处理匹配结果
            if not match_info:
                unmatched_fields.append(field_name)
                continue
            
            row, is_special = match_info
            
            # 填充数据
            if is_special and SPECIAL_PREFIX_MERGE_ROWS > 1:
                # 特殊前缀：合并多行填充
                merge_end_row = row + SPECIAL_PREFIX_MERGE_ROWS - 1
                
                # 先合并单元格（如果还未合并）
                try:
                    self.sheet.merge_cells(
                        start_row=row,
                        start_column=target_col,
                        end_row=merge_end_row,
                        end_column=target_col
                    )
                except ValueError:
                    # 已经合并，忽略
                    pass
                
                # 填充值到主单元格
                self.sheet.cell(row, target_col, value=field_value)
                matched_count[field_name] = matched_count.get(field_name, 0) + 1
            else:
                # 普通匹配：直接填充
                self.sheet.cell(row, target_col, value=field_value)
                matched_count[field_name] = matched_count.get(field_name, 0) + 1
        
        # 记录特殊前缀B列匹配失败的告警
        if special_prefix_no_match:
            unique_warnings = {}
            for row, a_col_val, field in special_prefix_no_match:
                key = (row, a_col_val)
                if key not in unique_warnings:
                    unique_warnings[key] = []
                unique_warnings[key].append(field)
            
            for (row, a_col_val), fields in unique_warnings.items():
                self.warnings.append(
                    f"⚠️  顶格表格特殊前缀B列匹配失败: 第{row}行 A列='{a_col_val}'，B列为空或不匹配字段 {fields}"
                )
        
        # 记录普通未匹配字段
        if unmatched_fields and SHOW_UNMATCHED_WARNINGS:
            section_name = log_section.get('name', '未知配置块')
            self.warnings.append(
                f"⚠️  顶格表格未匹配字段 ({section_name}): {unmatched_fields}"
            )
            self._suggest_field_mapping(unmatched_fields, start_row, end_row, is_sub_table=False)
        
        return matched_count
    
    def _suggest_field_mapping(self, unmatched_fields, start_row, end_row, is_sub_table=False):
        """
        为未匹配的字段提供映射建议
        检查Excel中的字段名，找出可能的匹配（部分匹配）
        """
        suggestions = []
        
        # 获取Excel中的所有字段名
        excel_fields = []
        # 顶格表和子表都只在B列搜索
        search_col = 2
        end_col = 2
        
        for row in range(start_row, end_row + 1):
            for col in range(search_col, end_col + 1):
                cell_value = self.sheet.cell(row, col).value
                if cell_value:
                    excel_fields.append(str(cell_value).strip())
        
        # 对每个未匹配的字段，查找可能的匹配
        for log_field in unmatched_fields:
            possible_matches = []
            for excel_field in excel_fields:
                # 检查是否有包含关系（子字符串）
                if log_field.lower() in excel_field.lower() or excel_field.lower() in log_field.lower():
                    possible_matches.append(excel_field)
            
            if possible_matches:
                suggestions.append(f"    '{log_field}' 可能对应: {possible_matches}")
        
        if suggestions:
            table_type = "子表" if is_sub_table else "顶格表格"
            self.warnings.append(f"💡 {table_type}字段映射建议（可在FIELD_NAME_MAPPING中配置）:")
            self.warnings.extend(suggestions)
    
    def match_and_fill_sub_table(self, log_section, start_row, end_row):
        """在子片段表格中匹配并填充数据（列1匹配）"""
        target_col = get_target_column()  # 使用配置的目标列
        matched_count = {}
        unmatched_fields = []
        
        for field_name, field_value in log_section['fields'].items():
            match_rows = []
            field_name_lower = field_name.lower()  # 转小写用于比较
            
            # 在B列（列1，因为A列是关键字）搜索字段名
            for row in range(start_row, end_row + 1):
                cell_value = self.sheet.cell(row, 2).value  # B列
                if cell_value:
                    cell_str = str(cell_value).strip()
                    cell_str_lower = cell_str.lower()  # 转小写比较
                    
                    # 精确匹配（不区分大小写）
                    if cell_str_lower == field_name_lower:
                        match_rows.append(row)
                    # 部分匹配（不区分大小写）
                    elif ENABLE_PARTIAL_MATCH and (
                        field_name_lower in cell_str_lower or 
                        cell_str_lower in field_name_lower
                    ):
                        match_rows.append(row)
            
            # 检查是否匹配
            if not match_rows:
                unmatched_fields.append(field_name)
                continue
            
            # 检查重复匹配
            if len(match_rows) > 1:
                self.warnings.append(
                    f"⚠️  子表重复匹配: 字段'{field_name}' 在第 {match_rows} 行都出现了"
                )
            
            # 填充数据
            for row in match_rows:
                self.sheet.cell(row, target_col, value=field_value)
                matched_count[field_name] = matched_count.get(field_name, 0) + 1
        
        # 记录未匹配字段
        if unmatched_fields and SHOW_UNMATCHED_WARNINGS:
            section_name = log_section.get('name', '未知配置块')
            self.warnings.append(
                f"⚠️  子表未匹配字段 ({section_name}): {unmatched_fields}"
            )
            # 检查是否有字段名映射建议
            self._suggest_field_mapping(unmatched_fields, start_row, end_row, is_sub_table=True)
        
        return matched_count
    
    def copy_sub_table(self, start_row, end_row, insert_after_row):
        """复制子片段表格到指定位置"""
        from openpyxl.cell.cell import MergedCell
        
        # 插入空行
        self.sheet.insert_rows(insert_after_row + 1, 1)
        insert_after_row += 1
        
        # 复制表格行
        table_rows = end_row - start_row + 1
        self.sheet.insert_rows(insert_after_row + 1, table_rows)
        
        # 复制数据和样式
        for offset in range(table_rows):
            source_row = start_row + offset
            target_row = insert_after_row + 1 + offset
            
            for col in range(1, self.sheet.max_column + 1):
                source_cell = self.sheet.cell(source_row, col)
                target_cell = self.sheet.cell(target_row, col)
                
                # 跳过MergedCell（合并单元格的非主单元格）
                if isinstance(target_cell, MergedCell):
                    continue
                
                # 复制值
                if not isinstance(source_cell, MergedCell):
                    target_cell.value = source_cell.value
                
                # 复制样式
                self.copy_cell_style(source_cell, target_cell)
        
        # 复制合并单元格区域
        for merged_range in list(self.sheet.merged_cells.ranges):
            # 检查是否在源表格范围内
            if merged_range.min_row >= start_row and merged_range.max_row <= end_row:
                # 计算目标位置的合并范围
                offset = insert_after_row + 1 - start_row
                new_min_row = merged_range.min_row + offset
                new_max_row = merged_range.max_row + offset
                new_min_col = merged_range.min_col
                new_max_col = merged_range.max_col
                
                # 添加新的合并单元格
                self.sheet.merge_cells(
                    start_row=new_min_row,
                    start_column=new_min_col,
                    end_row=new_max_row,
                    end_column=new_max_col
                )
        
        return insert_after_row + table_rows
    
    def save(self, output_file):
        """保存Excel文件（增强版，带完整验证）"""
        import os
        
        print(f"\n{'='*60}")
        print("[保存文件]")
        print(f"{'='*60}")
        
        # 1. 规范化路径
        output_file = os.path.normpath(output_file)
        abs_path = os.path.abspath(output_file)
        
        print(f"目标文件: {output_file}")
        print(f"完整路径: {abs_path}")
        print(f"当前目录: {os.getcwd()}")
        
        # 2. 检查并创建目录
        output_dir = os.path.dirname(abs_path)
        if output_dir:
            if not os.path.exists(output_dir):
                try:
                    os.makedirs(output_dir)
                    print(f"✓ 创建目录: {output_dir}")
                except Exception as e:
                    print(f"✗ 创建目录失败: {e}")
                    return
            else:
                print(f"✓ 目录已存在: {output_dir}")
        
        # 3. 检查文件是否被占用
        if os.path.exists(abs_path):
            try:
                # 尝试重命名检测占用
                import time
                temp_name = abs_path + '.tmp_test'
                os.rename(abs_path, temp_name)
                os.rename(temp_name, abs_path)
                print(f"✓ 文件可以覆盖")
            except OSError:
                print(f"⚠️  文件可能被占用，使用新文件名")
                timestamp = time.strftime("%Y%m%d_%H%M%S")
                base, ext = os.path.splitext(abs_path)
                abs_path = f"{base}_{timestamp}{ext}"
                print(f"新文件名: {abs_path}")
        
        # 4. 保存文件
        try:
            print(f"正在保存...")
            self.wb.save(abs_path)
            
            # 5. 验证文件已创建
            if os.path.exists(abs_path):
                file_size = os.path.getsize(abs_path)
                print(f"\n{'='*60}")
                print("✅ 保存成功!")
                print(f"{'='*60}")
                print(f"文件位置: {abs_path}")
                print(f"文件大小: {file_size:,} 字节")
                
                # 尝试在资源管理器中显示（仅Windows）
                if sys.platform == 'win32':
                    try:
                        import subprocess
                        subprocess.Popen(f'explorer /select,"{abs_path}"')
                        print(f"✓ 已在资源管理器中打开文件位置")
                    except:
                        pass
                
                print(f"{'='*60}\n")
                
            else:
                print(f"\n{'='*60}")
                print("✗ 保存失败!")
                print(f"{'='*60}")
                print(f"错误: 文件不存在于预期位置")
                print(f"请检查:")
                print(f"  1. 是否有目录写入权限")
                print(f"  2. 磁盘空间是否充足")
                print(f"  3. 路径是否正确")
                print(f"{'='*60}\n")
                
        except PermissionError as e:
            print(f"\n{'='*60}")
            print("✗ 权限错误!")
            print(f"{'='*60}")
            print(f"错误: {e}")
            print(f"解决方案:")
            print(f"  1. 以管理员权限运行脚本")
            print(f"  2. 选择其他有权限的目录")
            print(f"  3. 关闭占用该文件的程序（如Excel）")
            print(f"{'='*60}\n")
            
        except Exception as e:
            print(f"\n{'='*60}")
            print("✗ 保存失败!")
            print(f"{'='*60}")
            print(f"错误: {e}")
            import traceback
            traceback.print_exc()
            print(f"{'='*60}\n")

