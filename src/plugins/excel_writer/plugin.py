"""
Excel写入插件 - 将配置信息写入Excel模板
"""

import os
import re

import yaml
from openpyxl import load_workbook

from src.plugins.base import Plugin
from src.plugins.excel_writer.processor import ExcelProcessor


from src.utils import info, error, warning


class ExcelWriterPlugin(Plugin):
    """Excel写入插件 - Level 3 (Processor)"""

    level = 3  # 处理层
    dependencies = [
        "dld_configtmp",
        "config_parser",
        "constraint_checker",
    ]  # 依赖模板下载、配置解析和约束检查插件

    def execute(self, context: dict) -> dict:
        """将配置信息写入Excel模板

        Args:
            context: 上下文字典，需要包含：
                - excel_file: Excel模板文件路径
                - output_file: 输出文件路径（可选）
                - sheet_name: 工作表名称（可选）
                - config_parser.sections: 配置块列表
                - config_parser.parser: LogParser实例

        Returns:
            {
                'output_file': 输出文件路径,
                'processor': ExcelProcessor实例（供小插件使用）,
                'workbook': openpyxl Workbook对象
            }
        """
        # 获取输入参数
        # 优先使用 dld_configtmp 下载的模板，否则使用传入的 excel_file
        dld_configtmp_data = context.get("dld_configtmp", {})
        excel_file = dld_configtmp_data.get("template_path") or context.get(
            "excel_file"
        )
        output_file = context.get("output_file")
        sheet_name = context.get("sheet_name")

        if not excel_file:
            raise ValueError(
                "excel_writer: context 中缺少 'excel_file' 或 'dld_configtmp.template_path'"
            )

        # 获取 config_parser 的输出
        config_data = context.get("config_parser", {})
        sections = config_data.get("sections", [])
        parser = config_data.get("parser")

        if not sections or not parser:
            raise ValueError("excel_writer: 需要 config_parser 的输出数据")

        info(f"[Excel写入] 加载模板: {excel_file}")
        # 检查是否有多个TopConfig
        enable_top_table = self.config.get("top_table", {}).get("enable", True)
        if enable_top_table:
            top_keyword = self.config.get("top_table", {}).get("log_keyword", "opSch")
            groups = parser.group_by_top_config(top_keyword)
            num_top_configs = len(groups)

            if num_top_configs > 1:
                # 多个 TopConfig：根据配置决定处理方式
                multi_top_mode = self.config.get("top_table", {}).get(
                    "multi_top_mode", "multi_sheets"
                )

                if multi_top_mode == "multi_sheets":
                    info(
                        f"[Excel写入] 检测到 {num_top_configs} 个 TopConfig，使用多工作表模式（分页签）"
                    )
                    return self._process_multi_sheets(
                        excel_file, output_file, sheet_name, groups
                    )

                # single_sheet mode
                info(
                    f"[Excel写入] 检测到 {num_top_configs} 个 TopConfig，使用单工作表模式（分表格）"
                )
                return self._process_single_sheet_multi_tops(
                    excel_file, output_file, sheet_name, sections, groups
                )
            elif num_top_configs == 1:
                info("[Excel写入] 检测到 1 个 TopConfig，使用单工作表模式")
            else:
                info("[Excel写入] 未检测到 TopConfig")
        # 单工作表模式（单个或无 TopConfig）
        return self._process_single_sheet(
            excel_file, output_file, sheet_name, sections, parser
        )

    def _process_single_sheet(
        self, excel_file, output_file, sheet_name, sections, parser
    ):
        """单工作表处理模式"""
        # 加载Excel
        processor = ExcelProcessor(excel_file, sheet_name, self.config)
        info(f"[Excel写入] 使用工作表: {processor.sheet.title}")
        # 处理顶格表格
        enable_top_table = self.config.get("top_table", {}).get("enable", True)
        if enable_top_table:
            top_keyword = self.config.get("top_table", {}).get("log_keyword", "opSch")
            top_section = parser.get_top_section(top_keyword)
            if top_section:
                start_row, end_row = processor.find_top_table()
                if start_row:
                    matched = processor.match_and_fill_top_table(
                        top_section, start_row, end_row
                    )
                    info(f"[Excel写入] ✓ 顶格表格: 匹配 {len(matched)} 个字段")
                else:
                    error("[Excel写入] ✗ 未找到顶格表格")
            else:
                error(f"[Excel写入] ✗ 日志中未找到 {top_keyword}")
        # 处理子表
        self._process_sub_sections(processor, sections)

        # 生成输出文件名
        if not output_file:
            base_name = os.path.splitext(excel_file)[0]
            output_file = f"{base_name}_processed.xlsx"

        # 保存文件
        processor.save(output_file)
        info(f"[Excel写入] ✓ 保存完成: {output_file}")
        return {"output_file": output_file, "processor": processor}

    def _process_single_sheet_multi_tops(
        self, excel_file, output_file, sheet_name, sections, groups
    ):
        """单工作表处理多个TopConfig模式（所有top表放在同一个工作表）"""
        # 加载Excel
        processor = ExcelProcessor(excel_file, sheet_name, self.config)
        info(f"[Excel写入] 使用工作表: {processor.sheet.title}")
        info(f"[Excel写入] 处理 {len(groups)} 个 TopConfig 组")
        # 处理第一个 TopConfig 组（使用原有顶格表格）
        enable_top_table = self.config.get("top_table", {}).get("enable", True)
        if enable_top_table and len(groups) > 0:
            first_group = groups[0]
            top_section = first_group["top"]

            start_row, end_row = processor.find_top_table()
            if start_row:
                matched = processor.match_and_fill_top_table(
                    top_section, start_row, end_row
                )
                info(f"[Excel写入] ✓ TopConfig #1: 顶格表格匹配 {len(matched)} 个字段")
            else:
                error("[Excel写入] ✗ 未找到顶格表格")
            # 如果有多个 TopConfig，显示警告
            if len(groups) > 1:
                warning(
                    f"[Excel写入] ⚠️  检测到 {len(groups)} 个 TopConfig，"
                    f"但单工作表模式下顶格表格只能显示第一个"
                )
                warning(
                    "[Excel写入] ⚠️  建议使用 multi_top_mode: 'multi_sheets' 以显示所有 TopConfig"
                )

        # 处理所有组的子表（合并到同一个工作表）
        self._process_sub_sections(processor, sections)

        # 生成输出文件名
        if not output_file:
            base_name = os.path.splitext(excel_file)[0]
            output_file = f"{base_name}_processed.xlsx"

        # 保存文件
        processor.save(output_file)
        info(f"[Excel写入] ✓ 保存完成: {output_file}")
        return {"output_file": output_file, "processor": processor}

    def _process_multi_sheets(self, excel_file, output_file, sheet_name, groups):
        """多工作表处理模式"""
        workbook, template_sheet, template_name = self._setup_multi_sheets_workbook(
            excel_file, sheet_name, len(groups)
        )

        created_sheets, first_processor = self._create_all_sheets(
            workbook, template_sheet, groups
        )

        output_file = self._finalize_multi_sheets(
            workbook,
            template_name,
            excel_file,
            output_file,
            created_sheets,
        )

        return {
            "output_file": output_file,
            "processor": first_processor,
            "workbook": workbook,
            "created_sheets": created_sheets,
        }

    def _setup_multi_sheets_workbook(self, excel_file, sheet_name, num_groups):
        """设置多工作表工作簿"""
        workbook = load_workbook(excel_file)

        if sheet_name and sheet_name in workbook.sheetnames:
            template_sheet = workbook[sheet_name]
        else:
            template_sheet = workbook.active

        template_name = template_sheet.title
        info(f"[Excel写入] 模板工作表: {template_name}")
        info(f"[Excel写入] 创建 {num_groups} 个工作表...")
        return workbook, template_sheet, template_name

    def _create_all_sheets(self, workbook, template_sheet, groups):
        """为所有组创建工作表"""
        created_sheets = []
        first_processor = None

        for idx, group in enumerate(groups, 1):
            processor, sheet_title = self._process_group_to_sheet(
                workbook, template_sheet, group, idx, len(groups)
            )
            created_sheets.append((sheet_title, processor))

            if idx == 1:
                first_processor = processor

        return created_sheets, first_processor

    def _process_group_to_sheet(self, workbook, template_sheet, group, idx, total):
        """处理单个组到工作表"""
        top_section = group["top"]
        sub_sections = group["subs"]

        sheet_title = self._generate_sheet_name(top_section, idx)
        new_sheet = workbook.copy_worksheet(template_sheet)
        new_sheet.title = sheet_title

        processor = ExcelProcessor(workbook, new_sheet, self.config)

        enable_top_table = self.config.get("top_table", {}).get("enable", True)
        if enable_top_table:
            start_row, end_row = processor.find_top_table()
            if start_row:
                matched = processor.match_and_fill_top_table(
                    top_section, start_row, end_row
                )
                info(f"  [{idx}/{total}] {sheet_title}: 顶格表格 {len(matched)} 字段")
        self._process_sub_sections(processor, sub_sections)
        info(f"  [{idx}/{total}] {sheet_title}: 处理了 {len(sub_sections)} 个子表")
        return processor, sheet_title

    def _finalize_multi_sheets(
        self,
        workbook,
        template_name,
        excel_file,
        output_file,
        created_sheets,
    ):
        """完成多工作表处理"""
        if template_name in workbook.sheetnames:
            del workbook[template_name]

        if not output_file:
            base_name = os.path.splitext(excel_file)[0]
            output_file = f"{base_name}_multi_sheets.xlsx"

        workbook.save(output_file)
        info(f"[Excel写入] ✓ 保存完成: {output_file}")
        info(f"[Excel写入] 工作表数量: {len(created_sheets)}")
        return output_file

    def _process_sub_sections(self, processor, sections):
        """处理子表部分"""
        keyword_mapping = self.config.get("keyword_mapping", {})
        top_keyword = self.config.get("top_table", {}).get("log_keyword", "opSch")
        merge_rows = self.config.get("special_prefix", {}).get("merge_rows", 2)

        keyword_info = self._scan_sub_table_positions(processor, keyword_mapping)
        self._fill_all_sections(
            processor, sections, keyword_info, keyword_mapping, top_keyword, merge_rows
        )
        self._cleanup_unused_sub_tables(processor, keyword_info)

    def _scan_sub_table_positions(self, processor, keyword_mapping):
        """扫描所有子表的原始位置

        支持占位符关键字（如 IN__x__Cfg），会自动扫描 IN0Cfg, IN1Cfg, IN2Cfg 等

        占位符语法：使用 __x__ 标识需要替换的位置
        - IN__x__Cfg → IN0Cfg, IN1Cfg, IN2Cfg ...
        - ExCfg-ER → ExCfg-ER (普通关键字，不替换)
        """
        keyword_info = {}
        for excel_keyword in keyword_mapping.keys():
            # 检查是否包含占位符 __x__
            if "__x__" in excel_keyword:
                # 扫描所有可能的索引实例（0-9）
                for idx in range(10):
                    expanded_keyword = excel_keyword.replace("__x__", str(idx))
                    start_row, end_row = processor.find_sub_table(expanded_keyword)
                    if start_row:
                        keyword_info[expanded_keyword] = {
                            "orig_start": start_row,
                            "orig_end": end_row,
                            "count": 0,
                            "used": False,
                            "template": excel_keyword,  # 记录原始模板
                            "index": idx,  # 记录索引
                        }
            else:
                # 普通关键字，直接查找
                start_row, end_row = processor.find_sub_table(excel_keyword)
                if start_row:
                    keyword_info[excel_keyword] = {
                        "orig_start": start_row,
                        "orig_end": end_row,
                        "count": 0,
                        "used": False,
                    }
        return keyword_info

    def _fill_all_sections(
        self,
        processor,
        sections,
        keyword_info,
        keyword_mapping,
        top_keyword,
        merge_rows,
    ):
        """填充所有配置块到子表（按日志中出现的顺序）"""
        global_last_row = processor.sheet.max_row

        # 按日志中 sections 的顺序遍历（保持日志原始顺序）
        for section in sections:
            if section["name"] == top_keyword:
                continue

            matched_keyword = self._find_matching_keyword(
                section["name"], keyword_mapping
            )
            if not matched_keyword or matched_keyword not in keyword_info:
                continue

            global_last_row = self._fill_section_to_sub_table(
                processor,
                section,
                keyword_info[matched_keyword],
                global_last_row,
                merge_rows,
            )

        return global_last_row

    def _find_matching_keyword(self, section_name, keyword_mapping):
        """查找匹配的Excel关键词

        支持占位符关键字（使用 __x__ 标识）：
        - 日志: InxCfg0 → Excel: IN0Cfg (配置中为 IN__x__Cfg)
        - 日志: InxCfg1 → Excel: IN1Cfg
        - 日志: ERCfg (grp = 0) → Excel: ExCfg-ER (普通关键字，不替换)

        占位符语法：__x__ 表示需要从日志中提取数字替换的位置
        """
        for excel_keyword, log_pattern in keyword_mapping.items():
            match = re.match(log_pattern, section_name)
            if match:
                # 检查是否包含占位符 __x__
                if "__x__" in excel_keyword:
                    # 从日志 section name 中提取数字索引
                    # 例如：InxCfg0 → 提取 0，InxCfg1 → 提取 1
                    digit_match = re.search(r"(\d+)", section_name)
                    if digit_match:
                        index = digit_match.group(1)
                        # 返回展开后的关键字：IN__x__Cfg → IN0Cfg
                        return excel_keyword.replace("__x__", index)
                    else:
                        # 如果没找到数字，使用 0
                        return excel_keyword.replace("__x__", "0")
                else:
                    # 普通关键字，直接返回
                    return excel_keyword
        return None

    def _fill_section_to_sub_table(
        self, processor, section, table_info, global_last_row, merge_rows
    ):
        """填充配置块到子表"""
        if table_info["count"] == 0:
            processor.match_and_fill_sub_table(
                section, table_info["orig_start"], table_info["orig_end"]
            )
            table_info["count"] += 1
            table_info["used"] = True
            return global_last_row

        new_start = global_last_row + merge_rows

        # 复制子表并获取实际的结束行（copy_sub_table返回最后一行）
        actual_end = processor.copy_sub_table(
            table_info["orig_start"], table_info["orig_end"], new_start
        )
        # 使用 new_start+1 因为子表从 insert_after_row+1 开始
        processor.match_and_fill_sub_table(section, new_start + 1, actual_end)
        table_info["count"] += 1

        return actual_end

    def _cleanup_unused_sub_tables(self, processor, keyword_info):
        """清理未使用的子表及其周围的空行"""
        rows_to_delete = [
            (info["orig_start"], info["orig_end"])
            for info in keyword_info.values()
            if not info["used"]
        ]

        if not rows_to_delete:
            return

        rows_to_delete.sort(reverse=True)
        for start_row, end_row in rows_to_delete:
            # 扩展删除范围，包括子表后的空行（最多2行）
            extended_end = end_row
            for check_row in range(
                end_row + 1, min(end_row + 3, processor.sheet.max_row + 1)
            ):
                if self._is_empty_row(processor, check_row):
                    extended_end = check_row
                else:
                    break

            num_rows = extended_end - start_row + 1
            processor.sheet.delete_rows(start_row, num_rows)

        processor.normalize_subtable_spacing()

    def _is_empty_row(self, processor, row):
        """检查一行是否为空"""
        for col in range(1, min(processor.sheet.max_column + 1, 10)):  # 检查前10列
            cell_value = processor.sheet.cell(row, col).value
            if cell_value is not None and str(cell_value).strip():
                return False
        return True

    def _generate_sheet_name(self, top_section, index):
        """生成工作表名称"""
        # 尝试加载 auto_filename 插件配置
        try:
            auto_filename_config_path = os.path.join(
                os.path.dirname(__file__), "..", "auto_filename", "config.yaml"
            )
            if os.path.exists(auto_filename_config_path):
                with open(auto_filename_config_path, "r", encoding="utf-8") as f:
                    auto_filename_config = yaml.safe_load(f) or {}
            else:
                auto_filename_config = {}
        except (FileNotFoundError, yaml.YAMLError, PermissionError) as e:
            from src.utils import warning

            warning(f"无法加载 auto_filename 配置: {e}")
            auto_filename_config = {}

        if auto_filename_config.get("enable", False):
            filename_fields = auto_filename_config.get("fields", [])
            filename_value_mapping = auto_filename_config.get("value_mapping", {})

            if filename_fields:
                name_parts = []
                for field_name in filename_fields[:2]:  # 只用前2个字段
                    if field_name in top_section["fields"]:
                        value = top_section["fields"][field_name]
                        # 应用映射
                        if field_name in filename_value_mapping:
                            field_mapping = filename_value_mapping[field_name]
                            value_str = str(value).strip()
                            if value_str in field_mapping:
                                value = field_mapping[value_str]
                        name_parts.append(str(value))

                if name_parts:
                    sheet_name = f"Config_{index}_{'_'.join(name_parts)}"
                    # Excel工作表名称限制：31字符
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:28] + "..."
                    return sheet_name

        # 默认使用序号
        return f"Config_{index}"
