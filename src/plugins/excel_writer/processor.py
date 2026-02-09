"""
Excel处理器 - 负责Excel表格的读取、匹配和填充
"""

import contextlib
import os
import sys
from copy import copy

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from src.plugins.base import get_target_column
from src.utils import error, info, warning


class ExcelProcessor:
    """处理Excel表格的读取、匹配和填充"""

    def __init__(self, excel_file_or_workbook, sheet_name_or_sheet=None, config=None):
        """
        初始化ExcelProcessor

        Args:
            excel_file_or_workbook: Excel文件路径 或 openpyxl Workbook对象
            sheet_name_or_sheet: 工作表名称 或 openpyxl Worksheet对象
            config: 配置字典
        """
        self.warnings = []  # 存储告警信息
        self.config = config or {}

        # 判断是文件路径还是Workbook对象
        if isinstance(excel_file_or_workbook, str):
            # 文件路径模式
            self.excel_file = excel_file_or_workbook
            self.wb = load_workbook(excel_file_or_workbook)
            self.sheet = (
                self.wb[sheet_name_or_sheet] if sheet_name_or_sheet else self.wb.active
            )
        else:
            # Workbook对象模式
            self.excel_file = None
            self.wb = excel_file_or_workbook
            self.sheet = sheet_name_or_sheet if sheet_name_or_sheet else self.wb.active

    def get_cell_value_smart(self, row, col):
        """智能获取单元格值，处理合并单元格

        如果是MergedCell，找到合并区域的主单元格并返回其值

        Args:
            row: 行号（从1开始）
            col: 列号（从1开始）

        Returns:
            单元格的值，如果是MergedCell则返回主单元格的值
        """
        cell = self.sheet.cell(row, col)

        # 如果是普通单元格，直接返回值
        if not isinstance(cell, MergedCell):
            return cell.value

        # 如果是MergedCell，找到对应的合并区域
        for merged_range in self.sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 获取合并区域的主单元格（左上角）
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                master_cell = self.sheet.cell(min_row, min_col)
                return master_cell.value

        # 如果没有找到合并区域（不应该发生），返回None
        return None

    def copy_cell_style(self, source_cell, target_cell):
        """复制单元格样式"""
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    def normalize_subtable_spacing(self):
        """规范化所有子表之间的空行"""
        keyword_mapping = self.config.get("keyword_mapping", {})
        subtable_positions = self._find_all_subtable_positions(keyword_mapping)

        if len(subtable_positions) <= 1:
            return

        self._adjust_all_subtable_gaps(subtable_positions)

    def _find_all_subtable_positions(self, keyword_mapping):
        """查找所有子表的位置"""
        subtable_positions = []
        row = 1

        while row <= self.sheet.max_row:
            cell_a = self.sheet.cell(row, 1).value
            if cell_a and any(keyword in str(cell_a) for keyword in keyword_mapping):
                start_row, end_row = self._find_subtable_end(row, keyword_mapping)
                subtable_positions.append((start_row, end_row))
                row = end_row + 1
            else:
                row += 1

        return subtable_positions

    def _find_subtable_end(self, start_row, keyword_mapping):
        """查找子表的结束行"""
        end_row = start_row

        for check_row in range(
            start_row + 1, min(start_row + 20, self.sheet.max_row + 1)
        ):
            cell_check_a = self.sheet.cell(check_row, 1).value

            if cell_check_a and any(
                keyword in str(cell_check_a) for keyword in keyword_mapping
            ):
                break

            if self._row_has_data(check_row):
                end_row = check_row

        return start_row, end_row

    def _row_has_data(self, row):
        """检查行是否有数据"""
        cell_b = self.sheet.cell(row, 2).value
        cell_c = self.sheet.cell(row, 3).value
        cell_d = self.sheet.cell(row, 4).value
        return cell_b or cell_c or cell_d

    def _adjust_all_subtable_gaps(self, subtable_positions):
        """调整所有子表之间的间隙"""
        for i in range(len(subtable_positions) - 1, 0, -1):
            _, prev_end = subtable_positions[i - 1]
            current_start, _ = subtable_positions[i]
            gap = current_start - prev_end - 1

            if gap < 1:
                self._insert_gap(prev_end, subtable_positions, i)
            elif gap > 1:
                self._remove_excess_gap(prev_end, gap, subtable_positions, i)

    def _insert_gap(self, prev_end, subtable_positions, from_index):
        """插入空行"""
        self.sheet.insert_rows(prev_end + 1, 1)
        info(f"    ✓ 在第{prev_end}行后插入空行")
        for j in range(from_index, len(subtable_positions)):
            old_start, old_end = subtable_positions[j]
            subtable_positions[j] = (old_start + 1, old_end + 1)

    def _remove_excess_gap(self, prev_end, gap, subtable_positions, from_index):
        """删除多余空行"""
        rows_to_delete = gap - 1
        delete_start = prev_end + 2

        self.sheet.delete_rows(delete_start, rows_to_delete)
        info(f"    ✓ 删除多余空行（第{delete_start}行开始，共{rows_to_delete}行）")
        for j in range(from_index, len(subtable_positions)):
            old_start, old_end = subtable_positions[j]
            subtable_positions[j] = (
                old_start - rows_to_delete,
                old_end - rows_to_delete,
            )

    def find_top_table(self):
        """查找顶格表格的起始和结束行（顶格表格从第1行开始，无需关键字）"""
        keyword_mapping = self.config.get("keyword_mapping", {})
        start_row = self.config.get("top_table", {}).get("start_row", 1)

        # 查找结束行（遇到空行或子片段关键字）
        end_row = start_row
        for row in range(start_row, self.sheet.max_row + 1):
            # 检查前3列是否都为空（空行）
            if all(self.sheet.cell(row, col).value is None for col in range(1, 4)):
                end_row = row - 1
                break

            # 检查是否遇到子片段关键字（A列）
            first_cell = self.sheet.cell(row, 1).value
            if first_cell and str(first_cell).strip() in keyword_mapping:
                end_row = row - 1
                break

            end_row = row

        # 确保至少有一行
        if end_row < start_row:
            return None, None

        return start_row, end_row

    def find_sub_table(self, keyword):
        """查找子片段表格的起始和结束行"""
        keyword_mapping = self.config.get("keyword_mapping", {})
        start_row = None

        # 查找关键字所在行
        for row in range(1, self.sheet.max_row + 1):
            cell_value = self.sheet.cell(row, 1).value
            if cell_value and str(cell_value).strip() == keyword:
                start_row = row
                break

        if not start_row:
            return None, None

        # 子表的结束行（遇到空行或新的关键字）
        end_row = start_row
        for row in range(start_row + 1, self.sheet.max_row + 1):
            # 检查A列和B列是否都为空
            if (
                self.sheet.cell(row, 1).value is None
                and self.sheet.cell(row, 2).value is None
            ):
                break
            # 检查是否遇到新的子片段关键字
            first_cell = self.sheet.cell(row, 1).value
            if first_cell and str(first_cell).strip() in keyword_mapping:
                break
            end_row = row

        return start_row, end_row

    def match_and_fill_top_table(self, log_section, start_row, end_row):
        """在顶格表格中匹配并填充数据"""
        target_col = get_target_column(self.config)
        special_prefix_merge_rows = self.config.get("special_prefix", {}).get(
            "merge_rows", 2
        )
        show_unmatched_warnings = self.config.get("matching", {}).get(
            "show_unmatched_warnings", True
        )

        matched_count = {}
        unmatched_fields = []
        special_prefix_no_match = []

        for field_name, field_value in log_section["fields"].items():
            match_info = self._find_match_row_top_table(
                field_name, start_row, end_row, special_prefix_no_match
            )

            if not match_info:
                unmatched_fields.append(field_name)
                continue

            row, is_special = match_info
            from src.plugins.excel_writer.data_models import CellFillContext

            fill_ctx = CellFillContext(
                row=row,
                col=target_col,
                value=field_value,
                is_special=is_special,
                merge_rows=special_prefix_merge_rows,
            )
            self._fill_cell_value(fill_ctx)
            matched_count[field_name] = matched_count.get(field_name, 0) + 1

        from src.plugins.excel_writer.data_models import TopTableWarningContext

        warning_ctx = TopTableWarningContext(
            special_prefix_no_match=special_prefix_no_match,
            unmatched_fields=unmatched_fields,
            show_warnings=show_unmatched_warnings,
            log_section=log_section,
            start_row=start_row,
            end_row=end_row,
        )
        self._record_top_table_warnings(warning_ctx)

        return matched_count

    def _find_match_row_top_table(
        self, field_name, start_row, end_row, special_prefix_no_match
    ):
        """查找字段在顶格表格中的匹配行"""
        enable_partial_match = self.config.get("matching", {}).get(
            "enable_partial_match", True
        )
        special_prefix_for_b_column = self.config.get("special_prefix", {}).get(
            "for_b_column", []
        )
        field_name_lower = field_name.lower()

        for row in range(start_row, end_row + 1):
            a_col_value = self.get_cell_value_smart(row, 1)
            if not a_col_value:
                continue

            a_col_str = str(a_col_value).strip()
            is_special_prefix = any(
                a_col_str.startswith(prefix) for prefix in special_prefix_for_b_column
            )

            if is_special_prefix:
                from src.plugins.excel_writer.data_models import BColumnMatchContext

                ctx = BColumnMatchContext(
                    row=row,
                    field_name=field_name,
                    field_name_lower=field_name_lower,
                    a_col_str=a_col_str,
                    enable_partial_match=enable_partial_match,
                    special_prefix_no_match=special_prefix_no_match,
                )
                result = self._try_match_b_column(ctx)
                if result.matched:
                    return (result.row, True)
            else:
                match = self._try_match_a_column(
                    row, a_col_str, field_name_lower, enable_partial_match
                )
                if match:
                    return match

        return None

    def _try_match_b_column(self, ctx):
        """尝试匹配B列（特殊前缀情况）

        Args:
            ctx: BColumnMatchContext 包含匹配所需的所有上下文信息

        Returns:
            MatchResult: 匹配结果
        """
        from src.plugins.excel_writer.data_models import MatchResult

        b_col_value = self.get_cell_value_smart(ctx.row, 2)

        if b_col_value:
            b_col_str_lower = str(b_col_value).strip().lower()
            if b_col_str_lower == ctx.field_name_lower:
                return MatchResult.success(row=ctx.row, col=2, method="b_column_exact")
            if ctx.enable_partial_match and (
                ctx.field_name_lower in b_col_str_lower
                or b_col_str_lower in ctx.field_name_lower
            ):
                return MatchResult.success(
                    row=ctx.row, col=2, method="b_column_partial", confidence=0.8
                )
        else:
            if ctx.row not in [info[0] for info in ctx.special_prefix_no_match]:
                ctx.special_prefix_no_match.append(
                    (ctx.row, ctx.a_col_str, ctx.field_name)
                )

        return MatchResult.failure()

    def _try_match_a_column(
        self, row, a_col_str, field_name_lower, enable_partial_match
    ):
        """尝试匹配A列（普通情况）"""
        a_col_str_lower = a_col_str.lower()

        if a_col_str_lower == field_name_lower:
            return (row, False)
        if enable_partial_match and (
            field_name_lower in a_col_str_lower or a_col_str_lower in field_name_lower
        ):
            return (row, False)

        return None

    def _match_field_in_column(self, ctx) -> list:
        """通用字段匹配方法 - 在指定列中查找字段名匹配的行

        Args:
            ctx: ColumnMatchContext 包含匹配所需的所有上下文信息

        Returns:
            list: 匹配的行号列表
        """
        match_rows = []
        field_name_lower = ctx.field_name.lower()

        for row in range(ctx.start_row, ctx.end_row + 1):
            cell_value = self.sheet.cell(row, ctx.column).value
            if not cell_value:
                continue

            cell_str = str(cell_value).strip()
            cell_str_lower = cell_str.lower()

            # 精确匹配（不区分大小写）
            if (
                cell_str_lower == field_name_lower
                or ctx.enable_partial_match
                and (
                    field_name_lower in cell_str_lower
                    or cell_str_lower in field_name_lower
                )
            ):
                match_rows.append(row)

        return match_rows

    def _fill_cell_value(self, ctx):
        """填充单元格值

        Args:
            ctx: CellFillContext 包含单元格填充所需的所有信息
        """
        if ctx.is_special and ctx.merge_rows > 1:
            merge_end_row = ctx.row + ctx.merge_rows - 1
            with contextlib.suppress(ValueError):
                self.sheet.merge_cells(
                    start_row=ctx.row,
                    start_column=ctx.col,
                    end_row=merge_end_row,
                    end_column=ctx.col,
                )

        self.sheet.cell(ctx.row, ctx.col, value=ctx.value)

    def _record_top_table_warnings(self, ctx):
        """记录顶格表格的警告信息

        Args:
            ctx: TopTableWarningContext 包含警告记录所需的所有上下文信息
        """
        if ctx.special_prefix_no_match:
            unique_warnings = {}
            for row, a_col_val, field in ctx.special_prefix_no_match:
                key = (row, a_col_val)
                if key not in unique_warnings:
                    unique_warnings[key] = []
                unique_warnings[key].append(field)

            for (row, a_col_val), fields in unique_warnings.items():
                self.warnings.append(
                    f"⚠️  顶格表格特殊前缀B列匹配失败: 第{row}行 A列='{a_col_val}'，"
                    f"B列为空或不匹配字段 {fields}"
                )

        if ctx.unmatched_fields and ctx.show_warnings:
            section_name = ctx.log_section.get("name", "未知配置块")
            self.warnings.append(
                f"⚠️  顶格表格未匹配字段 ({section_name}): {ctx.unmatched_fields}"
            )
            self._suggest_field_mapping(
                ctx.unmatched_fields, ctx.start_row, ctx.end_row, is_sub_table=False
            )

    def _suggest_field_mapping(
        self, unmatched_fields, start_row, end_row, is_sub_table=False
    ):
        """
        为未匹配的字段提供映射建议
        检查Excel中的字段名，找出可能的匹配（部分匹配）
        """
        suggestions = []

        # 获取Excel中的所有字段名
        excel_fields = []
        # 顶格表和子表都只在B列搜索
        search_col = 2
        end_col = 2

        for row in range(start_row, end_row + 1):
            for col in range(search_col, end_col + 1):
                cell_value = self.sheet.cell(row, col).value
                if cell_value:
                    excel_fields.append(str(cell_value).strip())

        # 对每个未匹配的字段，查找可能的匹配
        for log_field in unmatched_fields:
            possible_matches = []
            for excel_field in excel_fields:
                # 检查是否有包含关系（子字符串）
                if (
                    log_field.lower() in excel_field.lower()
                    or excel_field.lower() in log_field.lower()
                ):
                    possible_matches.append(excel_field)

            if possible_matches:
                suggestions.append(f"    '{log_field}' 可能对应: {possible_matches}")

        if suggestions:
            table_type = "子表" if is_sub_table else "顶格表格"
            self.warnings.append(
                f"💡 {table_type}字段映射建议（可在FIELD_NAME_MAPPING中配置）:"
            )
            self.warnings.extend(suggestions)

    def match_and_fill_sub_table(self, log_section, start_row, end_row):
        """在子片段表格中匹配并填充数据（列1匹配）"""
        target_col = get_target_column(self.config)
        enable_partial_match = self.config.get("matching", {}).get(
            "enable_partial_match", True
        )
        show_unmatched_warnings = self.config.get("matching", {}).get(
            "show_unmatched_warnings", True
        )

        matched_count = {}
        unmatched_fields = []

        for field_name, field_value in log_section["fields"].items():
            # 使用通用字段匹配方法在B列查找
            from src.plugins.excel_writer.data_models import ColumnMatchContext

            col_match_ctx = ColumnMatchContext(
                field_name=field_name,
                start_row=start_row,
                end_row=end_row,
                column=2,
                enable_partial_match=enable_partial_match,
            )
            match_rows = self._match_field_in_column(col_match_ctx)

            # 检查是否匹配
            if not match_rows:
                unmatched_fields.append(field_name)
                continue

            # 检查重复匹配
            if len(match_rows) > 1:
                self.warnings.append(
                    f"⚠️  子表重复匹配: 字段'{field_name}' 在第 {match_rows} 行都出现了"
                )

            # 填充数据
            for row in match_rows:
                self.sheet.cell(row, target_col, value=field_value)
                matched_count[field_name] = matched_count.get(field_name, 0) + 1

        # 记录未匹配字段
        if unmatched_fields and show_unmatched_warnings:
            section_name = log_section.get("name", "未知配置块")
            self.warnings.append(
                f"⚠️  子表未匹配字段 ({section_name}): {unmatched_fields}"
            )
            # 检查是否有字段名映射建议
            self._suggest_field_mapping(
                unmatched_fields, start_row, end_row, is_sub_table=True
            )

        return matched_count

    def copy_sub_table(self, start_row, end_row, insert_after_row):
        """复制子片段表格到指定位置

        Args:
            start_row: 源子表起始行
            end_row: 源子表结束行
            insert_after_row: 在此行之后插入（新子表将从 insert_after_row+1 行开始）

        Returns:
            新子表的结束行号
        """
        # 计算子表行数并插入行
        table_rows = end_row - start_row + 1
        # 直接在指定位置插入行，不额外插入空行（merge_rows已处理间隙）
        self.sheet.insert_rows(insert_after_row + 1, table_rows)

        # 复制数据和样式
        for offset in range(table_rows):
            source_row = start_row + offset
            target_row = insert_after_row + 1 + offset

            for col in range(1, self.sheet.max_column + 1):
                source_cell = self.sheet.cell(source_row, col)
                target_cell = self.sheet.cell(target_row, col)

                # 跳过MergedCell（合并单元格的非主单元格）
                if isinstance(target_cell, MergedCell):
                    continue

                # 复制值
                if not isinstance(source_cell, MergedCell):
                    target_cell.value = source_cell.value

                # 复制样式
                self.copy_cell_style(source_cell, target_cell)

        # 复制合并单元格区域
        for merged_range in list(self.sheet.merged_cells.ranges):
            # 检查是否在源表格范围内
            if merged_range.min_row >= start_row and merged_range.max_row <= end_row:
                # 计算目标位置的合并范围
                offset = insert_after_row + 1 - start_row
                new_min_row = merged_range.min_row + offset
                new_max_row = merged_range.max_row + offset
                new_min_col = merged_range.min_col
                new_max_col = merged_range.max_col

                # 添加新的合并单元格
                self.sheet.merge_cells(
                    start_row=new_min_row,
                    start_column=new_min_col,
                    end_row=new_max_row,
                    end_column=new_max_col,
                )

        return insert_after_row + table_rows

    def save(self, output_file):
        """保存Excel文件（增强版，带完整验证）"""
        info(f"\n{'=' * 60}")
        info("[保存文件]")
        info(f"{'=' * 60}")
        abs_path = self._prepare_output_path(output_file)
        if not self._ensure_output_directory(abs_path):
            return

        abs_path = self._handle_file_conflict(abs_path)
        self._perform_save(abs_path)

    def _prepare_output_path(self, output_file):
        """规范化并准备输出路径"""
        output_file = os.path.normpath(output_file)
        abs_path = os.path.abspath(output_file)

        info(f"目标文件: {output_file}")
        info(f"完整路径: {abs_path}")
        info(f"当前目录: {os.getcwd()}")
        return abs_path

    def _ensure_output_directory(self, abs_path):
        """确保输出目录存在"""
        output_dir = os.path.dirname(abs_path)
        if not output_dir:
            return True

        if os.path.exists(output_dir):
            info(f"✓ 目录已存在: {output_dir}")
            return True

        try:
            os.makedirs(output_dir)
            info(f"✓ 创建目录: {output_dir}")
            return True
        except Exception as e:
            error(f"✗ 创建目录失败: {e}")
            return False

    def _handle_file_conflict(self, abs_path):
        """处理文件占用冲突"""
        if not os.path.exists(abs_path):
            return abs_path

        try:
            temp_name = abs_path + ".tmp_test"
            os.rename(abs_path, temp_name)
            os.rename(temp_name, abs_path)
            info("✓ 文件可以覆盖")
            return abs_path
        except OSError:
            warning("⚠️  文件可能被占用，使用新文件名")
            import time

            timestamp = time.strftime("%Y%m%d_%H%M%S")
            base, ext = os.path.splitext(abs_path)
            new_path = f"{base}_{timestamp}{ext}"
            info(f"新文件名: {new_path}")
            return new_path

    def _perform_save(self, abs_path):
        """执行保存并处理错误"""
        try:
            info("正在保存...")
            self.wb.save(abs_path)
            self._verify_and_report_success(abs_path)
        except PermissionError as e:
            self._print_permission_error(e)
        except Exception as e:
            self._print_general_error(e)

    def _verify_and_report_success(self, abs_path):
        """验证保存成功并报告"""
        if not os.path.exists(abs_path):
            self._print_file_not_found_error()
            return

        file_size = os.path.getsize(abs_path)
        info(f"\n{'=' * 60}")
        info("✅ 保存成功!")
        info(f"{'=' * 60}")
        info(f"文件位置: {abs_path}")
        info(f"文件大小: {file_size:,} 字节")
        self._try_open_in_explorer(abs_path)
        info(f"{'=' * 60}\n")

    def _try_open_in_explorer(self, abs_path):
        """尝试在资源管理器中打开（仅Windows）"""
        if sys.platform == "win32":
            try:
                import subprocess

                # 使用列表形式避免命令注入风险
                subprocess.Popen(["explorer", "/select,", abs_path])
                info("✓ 已在资源管理器中打开文件位置")
            except (OSError, subprocess.SubprocessError) as e:
                warning(f"无法在资源管理器中打开文件: {e}")

    def _print_file_not_found_error(self):
        """打印文件未找到错误"""
        info(f"\n{'=' * 60}")
        error("✗ 保存失败!")
        info(f"{'=' * 60}")
        error("错误: 文件不存在于预期位置")
        info("请检查:")
        info("  1. 是否有目录写入权限")
        info("  2. 磁盘空间是否充足")
        info("  3. 路径是否正确")
        info(f"{'=' * 60}\n")

    def _print_permission_error(self, err):
        """打印权限错误"""
        info(f"\n{'=' * 60}")
        error("✗ 权限错误!")
        info(f"{'=' * 60}")
        error(f"错误: {err}")
        info("解决方案:")
        info("  1. 以管理员权限运行脚本")
        info("  2. 选择其他有权限的目录")
        info("  3. 关闭占用该文件的程序（如Excel）")
        info(f"{'=' * 60}\n")

    def _print_general_error(self, err):
        """打印一般错误"""
        info(f"\n{'=' * 60}")
        error("✗ 保存失败!")
        info(f"{'=' * 60}")
        error(f"错误: {err}")
        import traceback

        traceback.print_exc()
        info(f"{'=' * 60}\n")
